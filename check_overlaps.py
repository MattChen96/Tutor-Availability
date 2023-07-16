import openpyxl
import json

with open('./data/caselle_excel_laboratori.json') as f:
    caselle = json.load(f)

def check_overlaps(file_path, lab_group):
    import openpyxl

    # Aprire il file di lavoro
    workbook = openpyxl.load_workbook(file_path)

    ws = workbook.worksheets[0]

    giorni_settimana = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì"]

    # Array contenente le righe di inizio dei vari laboratori
    arr_lab = list(caselle[lab_group].values())

    report = ""  # Stringa di report

    for j in range(len(arr_lab)):
        for k in range(j + 1, len(arr_lab)):
            for i in range(arr_lab[j], arr_lab[j] + 20):
                for l in range(4, 9):  # Colonne D a H
                    if (
                        ws.cell(row=i, column=l).value == ws.cell(
                            row=arr_lab[k] + (i - arr_lab[j]), column=l
                        ).value
                        and ws.cell(row=i, column=l).value is not None
                    ):
                        lab1 = ""
                        lab2 = ""
                        for lab, start_row in caselle[lab_group].items():
                            if start_row <= i <= start_row + 19:
                                lab1 = lab
                            if (
                                arr_lab[k] + (i - arr_lab[j])
                                >= start_row
                                and arr_lab[k] + (i - arr_lab[j])
                                <= start_row + 19
                            ):
                                lab2 = lab

                        report += f"Errore: Cognome '{ws.cell(row=i, column=l).value}' duplicato rilevato nelle righe {i} e {arr_lab[k] + (i - arr_lab[j])}, giorno: {giorni_settimana[l - 4]}, Laboratorio: {lab1} e {lab2}.\n"

    if not report:
        report = "Nessun cognome duplicato rilevato."

    print(report)

    # Salva il report in un file di testo
    report_file = "report.txt"
    with open(report_file, "w") as f:
        f.write(report)
    return report_file


