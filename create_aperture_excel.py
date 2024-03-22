import openpyxl
from openpyxl import Workbook
import json
import paths
import os

root_path = paths.ROOT_DIR
tmp_path = paths.TMP_DIR
data_path = paths.DATA_DIR

def create_aperture_excel(file_path, lab_group):

    print("START create_aperture_excel")

    with open( os.path.join(root_path, data_path, 'caselle_excel_laboratori.json')) as f:
        caselle = json.load(f)

    # Aprire il file di lavoro
    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.worksheets[0]

    #giorni_settimana = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì"]

    # Crea un nuovo foglio di lavoro
    wb = Workbook()
    ws_aperture = wb.active

    # Array contenente le righe di inizio dei vari laboratori
    tuple_lab = list(caselle[lab_group].items())

    for i in [2,3]:  #cicli per scrivere le fasce orarie -> copio dal excel originale
        for j in range(tuple_lab[0][1], tuple_lab[0][1] + 20):
            ws_aperture.cell(row=j, column=i, value=ws.cell(row=j, column=i).value)

    #Scansiono tutte le tabelle non appena trovo un apertura inserisco il nome
    #nella casella corrispondente della prima tabella --> in sostanza faccio una sovrapposizione di tutte le tabelle
    for lab_name, lab_index in tuple_lab:
        for l in range(4, 9):  # Colonne D a H
            indice = tuple_lab[0][1]
            for i in range(lab_index, lab_index + 20):
                #effettuo 3 controlli:
                #1. la cella non è vuota,
                #2. è presente una scritta
                #3.la casella non è colorata (il codice colore non inizia con FF) oppure è bianca (codice colore FFFFFFFF/0/00000000)
                if (ws.cell(row=i, column=l).value is not None
                            and ws.cell(row=i, column=l).value != ""
                            and (ws.cell(row=i, column=l).fill.start_color.index in ['FFFFFFFF', 0 , 00000000] or (not ws.cell(row=i, column=l).fill.start_color.index.startswith('FF')) )):

                    if ws_aperture.cell(row=indice, column=l).value is None:    #se la casella è vuota inserisco il nome
                        lista_lab = lab_name[3:]
                    else:                                                       #se la casella c'era già un nome lo concateno
                        lista_lab = ws_aperture.cell(row=indice, column=l).value + "/" + lab_name[3:]

                    ws_aperture.cell(row=indice, column=l, value=lista_lab)

                indice+=1

    file_name = f"aperture_{lab_group}.xlsx"
    excel_file = os.path.join(root_path, tmp_path, file_name)
    wb.save(excel_file)

    return excel_file